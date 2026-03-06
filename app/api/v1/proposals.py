"""
Proposals API router — /api/proposals/*
"""

import csv
import io

import openpyxl
from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, UploadFile

from app.api.deps import get_db
from app.schemas.proposal import ProposalCreate, ProposalResponse
from app.services.proposal_service import ProposalService

router = APIRouter(prefix="/api/proposals", tags=["Proposals"])


@router.post("/create", response_model=ProposalResponse)
async def create_proposal(
    request: ProposalCreate,
    background_tasks: BackgroundTasks,
    db=Depends(get_db),
):
    """Create a new proposal with automatic pricing calculation."""
    return ProposalService(db).create_proposal(request)


@router.get("")
def list_proposals(db=Depends(get_db)):
    """List all proposals."""
    return ProposalService(db).list_proposals()


@router.get("/{proposal_id}")
def get_proposal(proposal_id: int, db=Depends(get_db)):
    """Get full details for a single proposal."""
    return ProposalService(db).get_proposal(proposal_id)


@router.delete("/{proposal_id}")
def delete_proposal(proposal_id: int, db=Depends(get_db)):
    """Permanently delete a proposal and all related data."""
    return ProposalService(db).delete_proposal(proposal_id)


@router.patch("/{proposal_id}")
def update_proposal(proposal_id: int, payload: dict, db=Depends(get_db)):
    """Partially update editable proposal fields (client_name, client_email, project_name, status)."""
    return ProposalService(db).update_proposal(proposal_id, payload)


@router.put("/{proposal_id}/recalculate", response_model=ProposalResponse)
async def recalculate_proposal(
    proposal_id: int,
    request: ProposalCreate,
    db=Depends(get_db),
):
    """Re-run full pricing calculation for an existing proposal and update all stored data."""
    return ProposalService(db).recalculate_proposal(proposal_id, request)


@router.post("/{proposal_id}/send-email")
async def send_proposal_email(proposal_id: int, db=Depends(get_db)):
    """Generate the proposal PDF and email it to the client."""
    success = ProposalService(db).send_proposal_via_email(proposal_id)
    if success:
        return {"message": "Email sent successfully"}
    else:
        raise HTTPException(status_code=500, detail="Failed to send email. Check SMTP settings.")


@router.post("/import-complexity")
async def import_complexity(
    file: UploadFile = File(...),
    db=Depends(get_db),
):
    """
    Import complexity distribution from a CSV or Excel file.

    CSV format  : columns Report_ID, Complexity_Score
    Excel format: column named 'ComplexityScore'
    """
    try:
        contents = await file.read()
        filename = file.filename.lower()
        complexity_distribution = {0: 0, 1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        data_rows = []

        if filename.endswith(".csv"):
            csv_content = contents.decode("utf-8")
            data_rows = list(csv.DictReader(io.StringIO(csv_content)))

        elif filename.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            sheet = wb.active

            header_row_idx = -1
            complexity_col_idx = -1

            _target_names = {
                "complexityscore", "complexity_score",
                "complexity score", "complexity", "score",
            }
            for i, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
                if not any(cell is not None for cell in row):
                    continue
                for j, cell in enumerate(row):
                    if cell and str(cell).strip().lower() in _target_names:
                        header_row_idx = i
                        complexity_col_idx = j
                        break
                if complexity_col_idx != -1:
                    break

            if complexity_col_idx == -1:
                raise HTTPException(
                    status_code=400,
                    detail="Excel must contain a 'ComplexityScore' column",
                )

            for row in sheet.iter_rows(min_row=header_row_idx + 2, values_only=True):
                if complexity_col_idx < len(row):
                    data_rows.append({"ComplexityScore": row[complexity_col_idx]})
        else:
            raise HTTPException(status_code=400, detail="File must be CSV or Excel")

        _possible_cols = [
            "ComplexityScore", "complexityscore", "Complexity_Score",
            "complexity_score", "Complexity", "complexity", "Score", "score",
        ]
        for row in data_rows:
            col = next((c for c in _possible_cols if c in row), None)
            if not col:
                continue
            try:
                val = row[col]
                if val is None:
                    continue
                score = int(float(str(val).strip()))
                if 0 <= score <= 5:
                    complexity_distribution[score] = complexity_distribution.get(score, 0) + 1
            except (ValueError, TypeError):
                continue

        final = {k: v for k, v in complexity_distribution.items() if v > 0}
        return {
            "complexity_distribution": final,
            "total_reports": sum(final.values()),
            "message": "Complexity distribution imported successfully",
        }

    except HTTPException:
        raise
    except Exception as exc:
        import traceback
        print(traceback.format_exc())
        raise HTTPException(status_code=400, detail=f"Error parsing file: {exc}") from exc
