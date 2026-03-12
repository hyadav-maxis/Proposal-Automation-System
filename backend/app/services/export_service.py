"""
Export Service — generates PDF and Excel proposal documents.

Migrated from export_utils.py; logic is identical but now lives
inside the services layer where it belongs.
"""

import io
import os
from datetime import datetime
from typing import Any, Dict, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

try:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False

from app.core.config import settings


# ── Logo helper ───────────────────────────────────────────────────────────────


def _resolve_logo(logo_path: Optional[str] = None) -> Optional[str]:
    """Return a valid path to the company logo, or None."""
    base = os.path.dirname(os.path.abspath(__file__))
    # Walk up three levels to reach the project root (backend/app/services → backend/app → backend → root)
    root = os.path.dirname(os.path.dirname(os.path.dirname(base)))
    path = logo_path or settings.LOGO_PATH or "static/company_logo.png"
    if not os.path.isabs(path):
        path = os.path.join(root, path)
    if os.path.isfile(path):
        return path
    for ext in (".png", ".jpg", ".jpeg"):
        p = os.path.join(root, "static", f"company_logo{ext}")
        if os.path.isfile(p):
            return p
    return None


# ── PDF ───────────────────────────────────────────────────────────────────────


class ExportService:
    """Generates PDF and Excel exports from proposal data."""

    def generate_pdf(
        self,
        proposal_data: Dict[str, Any],
        output_path: Optional[str] = None,
        logo_path: Optional[str] = None,
    ) -> bytes:
        """Return PDF bytes (and optionally save to file)."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []

        logo_file = _resolve_logo(logo_path)
        if logo_file:
            try:
                img = RLImage(logo_file, width=2 * inch, height=0.75 * inch)
                story.append(img)
                story.append(Spacer(1, 0.2 * inch))
            except Exception:
                pass

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1a1a1a"),
            spaceAfter=30,
        )

        story.append(Paragraph("PROPOSAL", title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Basic Info
        info_data = [
            [
                Paragraph(f"<b>Proposal Number:</b> {proposal_data.get('proposal_number', '')}", styles["Normal"]),
                Paragraph(f"<b>Date:</b> {datetime.now().strftime('%B %d, %Y')}", styles["Normal"])
            ],
            [
                Paragraph(f"<b>Client:</b> {proposal_data.get('client_name', '')}", styles["Normal"]),
                Paragraph(f"<b>Project:</b> {proposal_data.get('project_name', '')}", styles["Normal"])
            ],
        ]
        
        # Extended Info
        db_size = proposal_data.get("database_size_gb")
        num_runs = proposal_data.get("number_of_runs")
        dep_type = (proposal_data.get("deployment_type") or "").replace("_", " ").title()
        res_loc = (proposal_data.get("resource_location") or "").replace("_", " ")
        
        meta_row1 = []
        if db_size is not None:
            meta_row1.append(Paragraph(f"<b>Database Size:</b> {db_size} GB", styles["Normal"]))
        if num_runs is not None:
            meta_row1.append(Paragraph(f"<b>Number of Runs:</b> {num_runs}", styles["Normal"]))
        
        meta_row2 = []
        if dep_type:
            meta_row2.append(Paragraph(f"<b>Deployment Type:</b> {dep_type}", styles["Normal"]))
        if res_loc:
            meta_row2.append(Paragraph(f"<b>Resource Location:</b> {res_loc}", styles["Normal"]))

        if meta_row1:
            if len(meta_row1) == 2:
                info_data.append(meta_row1)
            else:
                info_data.append([meta_row1[0], ""])
        if meta_row2:
            if len(meta_row2) == 2:
                info_data.append(meta_row2)
            else:
                info_data.append([meta_row2[0], ""])

        # Use more width to avoid overlap (letter size is 8.5" wide, ~6.5" usable)
        info_table = Table(info_data, colWidths=[3.5 * inch, 3.0 * inch])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3 * inch))

        # Pricing Breakdown
        story.append(Paragraph("<b>Pricing Breakdown</b>", styles["Heading2"]))
        pricing = proposal_data.get("pricing_breakdown", {})
        
        table_data = [["Item", "Quantity", "Total"]]
        
        skip_keys = {"subtotal", "total_price", "birt_complexity_breakdown"}
        
        # Sort keys to have a consistent order
        for key in sorted(pricing.keys()):
            if key in skip_keys:
                continue
            
            comp = pricing[key]
            if not isinstance(comp, dict):
                continue
            
            if key == "us_resources_surcharge":
                continue

            name = comp.get("name", key.replace("_", " ").title())
            qty = comp.get("quantity", 1)
            total = comp.get("total_price", 0)
            
            if total > 0:
                # Format quantity: 10.0 -> "10", 10.5 -> "10.5"
                qty_str = f"{qty:g}" if isinstance(qty, (int, float)) else str(qty)
                table_data.append([name, qty_str, f"${total:,.2f}"])

        # Subtotal
        # table_data.append([
        #     Paragraph("<b>Subtotal</b>", styles["Normal"]),
        #     "",
        #     Paragraph(f"<b>${pricing.get('subtotal', 0):,.2f}</b>", styles["Normal"])
        # ])
        
        # Surcharge
        # surcharge = pricing.get("us_resources_surcharge")
        # if surcharge and surcharge.get("total_price", 0) > 0:
        #     name = surcharge.get("name", "US-Based Resources Surcharge")
        #     table_data.append([name, "1", f"${surcharge['total_price']:,.2f}"])

        # Grand Total
        table_data.append([
            Paragraph("<b>TOTAL</b>", styles["Normal"]),
            "",
            Paragraph(f"<b>${pricing.get('total_price', 0):,.2f}</b>", styles["Normal"]),
        ])

        table = Table(table_data, colWidths=[3.5 * inch, 1 * inch, 1.5 * inch])
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 1), (0, -1), "LEFT"),     # Item column left
                    ("ALIGN", (1, 1), (1, -1), "CENTER"),   # Qty column center
                    ("ALIGN", (2, 1), (2, -1), "RIGHT"),    # Total column right
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(table)
        doc.build(story)

        pdf_bytes = buffer.getvalue()
        if output_path:
            with open(output_path, "wb") as f:
                f.write(pdf_bytes)
        return pdf_bytes

    # ── Excel ─────────────────────────────────────────────────────────────────

    def generate_excel(
        self,
        proposal_data: Dict[str, Any],
        logo_path: Optional[str] = None,
    ) -> bytes:
        """Return Excel bytes."""
        if not _OPENPYXL:
            raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

        wb = Workbook()
        ws = wb.active
        ws.title = "Proposal"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        bold_font = Font(bold=True)

        row_offset = 1
        logo_file = _resolve_logo(logo_path)
        if logo_file:
            try:
                img = XLImage(logo_file)
                img.width = 180
                img.height = 60
                ws.add_image(img, "A1")
                row_offset = 5
            except Exception:
                pass

        ws.cell(row=row_offset, column=1, value="PROPOSAL").font = Font(bold=True, size=16)
        row_offset += 1
        ws.cell(row=row_offset, column=1, value=f"Proposal Number: {proposal_data.get('proposal_number', '')}")
        row_offset += 1
        ws.cell(row=row_offset, column=1, value=f"Date: {datetime.now().strftime('%Y-%m-%d')}")
        row_offset += 1
        ws.cell(row=row_offset, column=1, value=f"Client: {proposal_data.get('client_name', '')}")
        row_offset += 1
        ws.cell(row=row_offset, column=1, value=f"Project: {proposal_data.get('project_name', '')}")
        row_offset += 1
        
        # Extended Metadata
        db_size = proposal_data.get("database_size_gb")
        num_runs = proposal_data.get("number_of_runs")
        dep_type = (proposal_data.get("deployment_type") or "").replace("_", " ").title()
        res_loc = (proposal_data.get("resource_location") or "").replace("_", " ")

        if db_size is not None:
            ws.cell(row=row_offset, column=1, value=f"Database Size: {db_size} GB")
            row_offset += 1
        if num_runs is not None:
            ws.cell(row=row_offset, column=1, value=f"Number of Runs: {num_runs}")
            row_offset += 1
        if dep_type:
            ws.cell(row=row_offset, column=1, value=f"Deployment Type: {dep_type}")
            row_offset += 1
        if res_loc:
            ws.cell(row=row_offset, column=1, value=f"Resource Location: {res_loc}")
            row_offset += 1

        row_offset += 1
        ws.cell(row=row_offset, column=1, value="Pricing Breakdown").font = bold_font
        row_offset += 1

        pricing = proposal_data.get("pricing_breakdown", {})
        rows = [("Item", "Quantity", "Amount")]
        
        skip_keys = {"subtotal", "total_price", "birt_complexity_breakdown"}
        for key in sorted(pricing.keys()):
            if key in skip_keys:
                continue
            
            comp = pricing[key]
            if not isinstance(comp, dict):
                continue
            
            if key == "us_resources_surcharge":
                continue
                
            name = comp.get("name", key.replace("_", " ").title())
            qty = comp.get("quantity", 1)
            total = comp.get("total_price", 0)
            if total > 0:
                rows.append((name, qty, total))

        rows.append(("Subtotal", "", pricing.get("subtotal", 0)))
        
        surcharge = pricing.get("us_resources_surcharge")
        if surcharge and surcharge.get("total_price", 0) > 0:
            name = surcharge.get("name", "USA-Based Resources Surcharge")
            rows.append((name, 1, surcharge["total_price"]))

        rows.append(("TOTAL", "", pricing.get("total_price", 0)))

        for i, (item, qty, amount) in enumerate(rows, start=row_offset):
            ws.cell(row=i, column=1, value=item)
            ws.cell(row=i, column=2, value=qty)
            if isinstance(amount, (int, float)):
                ws.cell(row=i, column=3, value=float(amount))
                ws.cell(row=i, column=3).number_format = '"$"#,##0.00'
            else:
                ws.cell(row=i, column=3, value=amount)
            
            if item == "Item":
                for col in (1, 2, 3):
                    ws.cell(row=i, column=col).fill = header_fill
                    ws.cell(row=i, column=col).font = header_font_white
            if item == "TOTAL":
                ws.cell(row=i, column=1).font = bold_font
                ws.cell(row=i, column=3).font = bold_font

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 15

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def generate_all_proposals_excel(self, proposals: list, filter_info: Optional[str] = None) -> bytes:
        """Export all proposals as a summary Excel table."""
        if not _OPENPYXL:
            raise RuntimeError("openpyxl is not installed")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "All Proposals"
        
        current_row = 1
        if filter_info:
            ws.cell(row=current_row, column=1, value=f"Filter: {filter_info}")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            ws.cell(row=current_row, column=1).font = Font(italic=True, color="666666")
            current_row += 1
        
        headers = ["Proposal No.", "Client Name", "Email", "Project", "Location", "Created At", "Status", "Amount"]
        for col_idx, text in enumerate(headers, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for p_idx, p in enumerate(proposals):
            row_idx = current_row + 1 + p_idx
            created = p.get("created_at")
            if isinstance(created, datetime):
                created = created.strftime("%Y-%m-%d %H:%M")
            
            loc = (p.get("resource_location") or "standard").replace("US_based", "US Based")
            
            ws.cell(row=row_idx, column=1, value=p.get("proposal_number"))
            ws.cell(row=row_idx, column=2, value=p.get("client_name"))
            ws.cell(row=row_idx, column=3, value=p.get("client_email"))
            ws.cell(row=row_idx, column=4, value=p.get("project_name"))
            ws.cell(row=row_idx, column=5, value=loc)
            ws.cell(row=row_idx, column=6, value=created)
            ws.cell(row=row_idx, column=7, value=p.get("status", "").upper())
            
            amount_cell = ws.cell(row=row_idx, column=8, value=float(p.get("total_price") or 0))
            amount_cell.number_format = '"$"#,##0.00'

        # Column widths
        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
            ws.column_dimensions[col_letter].width = 20
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 30
            
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def generate_all_proposals_pdf(self, proposals: list, logo_path: Optional[str] = None, filter_info: Optional[str] = None) -> bytes:
        """Export all proposals as a summary PDF table."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=0.5*inch, rightMargin=0.5*inch)
        story = []
        
        logo_file = _resolve_logo(logo_path)
        if logo_file:
            try:
                img = RLImage(logo_file, width=1.5*inch, height=0.5*inch)
                img.hAlign = 'LEFT'
                story.append(img)
            except Exception: pass
            
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("Title", parent=styles["Heading1"], fontSize=18, spaceAfter=10)
        story.append(Paragraph("All Proposals Report", title_style))
        
        if filter_info:
            filter_style = ParagraphStyle("Filter", parent=styles["Normal"], fontSize=10, textColor=colors.grey, spaceAfter=10)
            story.append(Paragraph(f"Filter: {filter_info}", filter_style))
            
        story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
        story.append(Spacer(1, 0.2*inch))
        
        header = ["Proposal No.", "Client", "Project", "Location", "Amount"]
        data = [header]
        
        for p in proposals:
            loc = (p.get("resource_location") or "standard").replace("US_based", "US Based")
            amount = f"${float(p.get('total_price') or 0):,.2f}"
            
            data.append([
                Paragraph(p.get("proposal_number", ""), styles["Normal"]),
                Paragraph(p.get("client_name", "") or "", styles["Normal"]),
                Paragraph(p.get("project_name", "") or "", styles["Normal"]),
                loc,
                amount
            ])
            
        table = Table(data, colWidths=[1.8*inch, 1.5*inch, 1.5*inch, 1.2*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        doc.build(story)
        return buffer.getvalue()

